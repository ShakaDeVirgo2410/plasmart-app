# =============================================================================
# PLASMART - GESTOR DE ÓRDENES DE TRABAJO v2.0
# plasmartcba.com | Córdoba, Argentina
#
# Instalación:
#   pip install streamlit pandas plotly openpyxl
#
# Ejecución:
#   streamlit run app.py
# =============================================================================

import streamlit as st
import sqlite3
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from datetime import date, datetime
import io
import os

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURACIÓN DE PÁGINA
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Plasmart · Gestor de OTs",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─────────────────────────────────────────────────────────────────────────────
# ESTILOS CSS
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&family=Syne:wght@700;800&display=swap');
    html, body, [class*="css"] { font-family: 'Inter', sans-serif; }
    .main .block-container { padding-top: 1.5rem; padding-bottom: 2rem; max-width: 1400px; }
    [data-testid="stSidebar"] { background: linear-gradient(180deg, #0f172a 0%, #1e293b 100%); }
    [data-testid="stSidebar"] * { color: #e2e8f0 !important; }
    [data-testid="stSidebar"] .stRadio label { font-size: 0.92rem !important; padding: 0.4rem 0.6rem !important; border-radius: 6px; transition: background 0.2s; }
    [data-testid="stSidebar"] .stRadio label:hover { background: rgba(255,255,255,0.08); }
    .kpi-card { background: white; border-radius: 12px; padding: 1.25rem 1.5rem; box-shadow: 0 1px 3px rgba(0,0,0,0.08); border-left: 4px solid #f97316; margin-bottom: 0.75rem; }
    .kpi-label { font-size: 0.75rem; font-weight: 600; text-transform: uppercase; letter-spacing: 0.05em; color: #64748b; margin-bottom: 0.4rem; }
    .kpi-value { font-size: 1.75rem; font-weight: 700; color: #0f172a; line-height: 1; font-family: 'Syne', sans-serif; }
    .kpi-sub { font-size: 0.78rem; color: #94a3b8; margin-top: 0.3rem; }
    .kpi-card.blue   { border-left-color: #3b82f6; }
    .kpi-card.green  { border-left-color: #22c55e; }
    .kpi-card.orange { border-left-color: #f97316; }
    .kpi-card.red    { border-left-color: #ef4444; }
    .kpi-card.purple { border-left-color: #a855f7; }
    .kpi-card.teal   { border-left-color: #14b8a6; }
    .section-title { font-family: 'Syne', sans-serif; font-size: 1.4rem; font-weight: 800; color: #0f172a; margin-bottom: 1rem; padding-bottom: 0.5rem; border-bottom: 2px solid #f1f5f9; }
    .logo-header { text-align: center; padding: 1.5rem 1rem 1rem; border-bottom: 1px solid rgba(255,255,255,0.1); margin-bottom: 1.5rem; }
    .logo-title { font-family: 'Syne', sans-serif; font-size: 1.5rem; font-weight: 800; color: #f97316 !important; letter-spacing: -0.02em; line-height: 1; }
    .logo-sub { font-size: 0.7rem; color: #94a3b8 !important; letter-spacing: 0.08em; text-transform: uppercase; margin-top: 3px; }
    .stButton > button { font-weight: 600; border-radius: 8px; transition: all 0.2s; }
    .stButton > button:hover { transform: translateY(-1px); box-shadow: 0 4px 12px rgba(0,0,0,0.15); }
    .info-box { background: #f0f9ff; border: 1px solid #bae6fd; border-radius: 10px; padding: 1rem 1.25rem; margin: 0.75rem 0; font-size: 0.88rem; color: #0c4a6e; }
    .factura-box { background: #f0fdf4; border: 1px solid #86efac; border-radius: 10px; padding: 1rem 1.25rem; margin: 0.75rem 0; font-size: 0.88rem; color: #166534; }
    .sin-factura-box { background: #fef9c3; border: 1px solid #fde047; border-radius: 10px; padding: 1rem 1.25rem; margin: 0.75rem 0; font-size: 0.88rem; color: #854d0e; }
    hr.thin { border: none; border-top: 1px solid #f1f5f9; margin: 1.25rem 0; }
    .estado { font-size: 11px; padding: 3px 9px; border-radius: 99px; font-weight: 600; }
    .e-anticipo { background:#fef3c7; color:#92400e; }
    .e-saldo    { background:#dbeafe; color:#1e40af; }
    .e-cerrada  { background:#dcfce7; color:#166534; }
</style>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────────────
# BASE DE DATOS
# ─────────────────────────────────────────────────────────────────────────────
DB_PATH = "plasmart.db"
EXCEL_PATH = "plasmart_ordenes.xlsx"

def get_conn():
    conn = sqlite3.connect(DB_PATH, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_conn()
    c = conn.cursor()

    # Tabla principal de OTs (con campos nuevos)
    c.execute("""
        CREATE TABLE IF NOT EXISTS ordenes (
            id_ot           TEXT PRIMARY KEY,
            cliente         TEXT NOT NULL,
            cuit            TEXT,
            fecha_pedido    TEXT NOT NULL,
            tipo_trabajo    TEXT,
            diseno          TEXT,
            monto_total     REAL NOT NULL,
            con_factura     INTEGER NOT NULL DEFAULT 0,
            iva_pct         REAL NOT NULL DEFAULT 21,
            monto_iva       REAL NOT NULL DEFAULT 0,
            monto_total_cf  REAL NOT NULL DEFAULT 0,
            pct_anticipo    REAL NOT NULL DEFAULT 50,
            monto_anticipo  REAL NOT NULL DEFAULT 0,
            fecha_anticipo  TEXT,
            kg_chapa        REAL NOT NULL DEFAULT 0,
            fecha_entrega   TEXT,
            fecha_saldo     TEXT,
            canal           TEXT,
            vendedor        TEXT,
            estado          TEXT NOT NULL DEFAULT 'Pendiente anticipo',
            notas           TEXT,
            archivada       INTEGER NOT NULL DEFAULT 0,
            created_at      TEXT NOT NULL DEFAULT (datetime('now','localtime'))
        )
    """)

    # Migración: agregar columnas nuevas si no existen (para bases de datos anteriores)
    columnas_nuevas = [
        ("cuit", "TEXT"),
        ("tipo_trabajo", "TEXT"),
        ("diseno", "TEXT"),
        ("con_factura", "INTEGER NOT NULL DEFAULT 0"),
        ("iva_pct", "REAL NOT NULL DEFAULT 21"),
        ("monto_iva", "REAL NOT NULL DEFAULT 0"),
        ("monto_total_cf", "REAL NOT NULL DEFAULT 0"),
    ]
    cols_existentes = [row[1] for row in c.execute("PRAGMA table_info(ordenes)").fetchall()]
    for col_name, col_type in columnas_nuevas:
        if col_name not in cols_existentes:
            c.execute(f"ALTER TABLE ordenes ADD COLUMN {col_name} {col_type}")

    # Tabla de clientes
    c.execute("""
        CREATE TABLE IF NOT EXISTS clientes (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre      TEXT NOT NULL UNIQUE,
            cuit        TEXT,
            telefono    TEXT,
            email       TEXT,
            created_at  TEXT NOT NULL DEFAULT (datetime('now','localtime'))
        )
    """)

    # Configuración mensual
    c.execute("""
        CREATE TABLE IF NOT EXISTS config_mensual (
            anio            INTEGER NOT NULL,
            mes             INTEGER NOT NULL,
            costo_chapa_kg  REAL DEFAULT 0,
            costo_mo        REAL DEFAULT 0,
            gastos_pub      REAL DEFAULT 0,
            PRIMARY KEY (anio, mes)
        )
    """)

    # Configuración global
    c.execute("""
        CREATE TABLE IF NOT EXISTS config_global (
            clave TEXT PRIMARY KEY,
            valor TEXT
        )
    """)

    defaults = [
        ("costo_chapa_kg_default", "800"),
        ("vendedores", "Martín,Lucía,Santiago,Valentina"),
        ("iva_default", "21"),
    ]
    for k, v in defaults:
        c.execute("INSERT OR IGNORE INTO config_global (clave, valor) VALUES (?,?)", (k, v))

    conn.commit()
    conn.close()

def next_ot_id():
    conn = get_conn()
    anio = date.today().year
    c = conn.cursor()
    c.execute("SELECT COUNT(*) FROM ordenes WHERE id_ot LIKE ?", (f"OT-{anio}-%",))
    n = c.fetchone()[0] + 1
    conn.close()
    return f"OT-{anio}-{n:04d}"

def calc_estado(fecha_anticipo, fecha_saldo):
    if fecha_saldo:
        return "Cerrada"
    if fecha_anticipo:
        return "Pendiente saldo"
    return "Pendiente anticipo"


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL AUTOMÁTICO
# ─────────────────────────────────────────────────────────────────────────────
def actualizar_excel():
    """Regenera el Excel completo cada vez que se modifica una OT."""
    conn = get_conn()
    df = pd.read_sql(
        "SELECT * FROM ordenes WHERE archivada=0 ORDER BY created_at DESC",
        conn
    )
    conn.close()

    if df.empty:
        return

    # Renombrar y ordenar columnas para el Excel
    col_map = {
        "id_ot":          "ID OT",
        "cliente":        "Cliente",
        "cuit":           "CUIT",
        "fecha_pedido":   "Fecha Pedido",
        "tipo_trabajo":   "Tipo de Trabajo",
        "diseno":         "Diseño",
        "canal":          "Canal de Venta",
        "vendedor":       "Vendedor",
        "kg_chapa":       "Kg Chapa",
        "monto_total":    "Monto Total (IPV)",
        "con_factura":    "Con Factura",
        "iva_pct":        "IVA %",
        "monto_iva":      "Monto IVA",
        "monto_total_cf": "Total con IVA",
        "pct_anticipo":   "% Anticipo",
        "monto_anticipo": "Monto Anticipo",
        "fecha_anticipo": "Fecha Anticipo",
        "fecha_entrega":  "Fecha Entrega",
        "fecha_saldo":    "Fecha Pago Saldo",
        "estado":         "Estado",
        "notas":          "Notas",
        "created_at":     "Fecha Carga",
    }

    cols_orden = list(col_map.keys())
    cols_presentes = [c for c in cols_orden if c in df.columns]
    df_excel = df[cols_presentes].rename(columns=col_map)

    # Convertir con_factura a texto legible
    if "Con Factura" in df_excel.columns:
        df_excel["Con Factura"] = df_excel["Con Factura"].apply(
            lambda x: "Sí" if x == 1 else "No"
        )

    # Escribir Excel con formato
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df_excel.to_excel(writer, index=False, sheet_name="Órdenes de Trabajo")

        ws = writer.sheets["Órdenes de Trabajo"]

        # Ancho de columnas automático
        for col in ws.columns:
            max_len = max(
                len(str(cell.value)) if cell.value else 0
                for cell in col
            )
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        # Estilo de encabezado
        from openpyxl.styles import PatternFill, Font, Alignment
        header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=10)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Altura del encabezado
        ws.row_dimensions[1].height = 22

        # Alternar colores de filas
        from openpyxl.styles import PatternFill as PF
        fill_par   = PF(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        fill_impar = PF(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
            fill = fill_par if i % 2 == 0 else fill_impar
            for cell in row:
                cell.fill = fill

    # Guardar en disco para descarga persistente
    with open(EXCEL_PATH, "wb") as f:
        f.write(buf.getvalue())

    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def fmt_ars(v):
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return "—"
    return f"$ {v:,.0f}".replace(",", ".")

def fmt_kg(v):
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return "—"
    return f"{v:,.1f} kg"

def to_excel_download(df: pd.DataFrame) -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        df.to_excel(w, index=False, sheet_name="Datos")
    return buf.getvalue()

def load_ordenes(archivadas=False) -> pd.DataFrame:
    conn = get_conn()
    arch = 1 if archivadas else 0
    df = pd.read_sql(
        "SELECT * FROM ordenes WHERE archivada=? ORDER BY created_at DESC",
        conn, params=(arch,)
    )
    conn.close()
    for col in ["fecha_pedido","fecha_anticipo","fecha_entrega","fecha_saldo"]:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors="coerce")
    return df

def load_clientes():
    conn = get_conn()
    df = pd.read_sql("SELECT * FROM clientes ORDER BY nombre", conn)
    conn.close()
    return df

def load_config_global():
    conn = get_conn()
    rows = conn.execute("SELECT clave,valor FROM config_global").fetchall()
    conn.close()
    return {r["clave"]: r["valor"] for r in rows}

def save_config_global(clave, valor):
    conn = get_conn()
    conn.execute("INSERT OR REPLACE INTO config_global(clave,valor) VALUES(?,?)", (clave, str(valor)))
    conn.commit()
    conn.close()

def load_config_mensual(anio, mes):
    conn = get_conn()
    row = conn.execute(
        "SELECT * FROM config_mensual WHERE anio=? AND mes=?", (anio, mes)
    ).fetchone()
    conn.close()
    if row:
        return dict(row)
    return {"anio": anio, "mes": mes, "costo_chapa_kg": 0.0, "costo_mo": 0.0, "gastos_pub": 0.0}

def save_config_mensual(anio, mes, costo_chapa_kg, costo_mo, gastos_pub):
    conn = get_conn()
    conn.execute("""
        INSERT OR REPLACE INTO config_mensual(anio,mes,costo_chapa_kg,costo_mo,gastos_pub)
        VALUES(?,?,?,?,?)
    """, (anio, mes, costo_chapa_kg, costo_mo, gastos_pub))
    conn.commit()
    conn.close()

CANALES       = ["Publicidad web", "Instagram/Redes", "Referencia", "Otros"]
ESTADOS       = ["Pendiente anticipo", "Pendiente saldo", "Cerrada"]
TIPOS_TRABAJO = ["Corte Láser", "Corte Láser y Plegado", "Plegado"]
DISENOS       = ["Cliente", "Plasmart"]
MESES_ES      = ["Enero","Febrero","Marzo","Abril","Mayo","Junio",
                 "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]


# ─────────────────────────────────────────────────────────────────────────────
# SIDEBAR
# ─────────────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("""
    <div class="logo-header">
        <div class="logo-title">⚡ PLASMART</div>
        <div class="logo-sub">Gestor de Órdenes</div>
    </div>
    """, unsafe_allow_html=True)

    pagina = st.radio(
        "Navegación",
        ["🏠 Dashboard", "📋 Órdenes de Trabajo", "👥 Clientes",
         "📊 Análisis Mensual", "📈 Análisis Anual", "⚙️ Configuración"],
        label_visibility="collapsed"
    )

    st.markdown("---")
    st.markdown(
        '<div style="font-size:0.72rem;color:#64748b;text-align:center;">plasmartcba.com<br>Córdoba, Argentina</div>',
        unsafe_allow_html=True
    )

# ─────────────────────────────────────────────────────────────────────────────
# INIT DB
# ─────────────────────────────────────────────────────────────────────────────
init_db()


# ─────────────────────────────────────────────────────────────────────────────
# ══════════════════════════════ DASHBOARD ════════════════════════════════════
# ─────────────────────────────────────────────────────────────────────────────
if pagina == "🏠 Dashboard":
    st.markdown('<div class="section-title">🏠 Dashboard Principal</div>', unsafe_allow_html=True)

    df_all = load_ordenes()
    hoy = date.today()
    mes_actual  = hoy.month
    anio_actual = hoy.year

    df_activas = df_all[df_all["estado"] != "Cerrada"]
    df_mes = df_all[
        (df_all["fecha_saldo"].dt.month == mes_actual) &
        (df_all["fecha_saldo"].dt.year  == anio_actual) &
        (df_all["estado"] == "Cerrada")
    ]

    cfg   = load_config_mensual(anio_actual, mes_actual)
    cfg_g = load_config_global()
    costo_kg_default = float(cfg_g.get("costo_chapa_kg_default", 800))

    ingresos_mes     = df_mes["monto_total"].sum() if not df_mes.empty else 0
    kg_mes           = df_mes["kg_chapa"].sum()    if not df_mes.empty else 0
    ordenes_abiertas = len(df_activas)
    costo_chapa_mes  = kg_mes * (cfg["costo_chapa_kg"] or costo_kg_default)
    costo_mo_mes     = cfg["costo_mo"] or 0
    comisiones_mes   = ingresos_mes * 0.01
    gastos_pub_mes   = cfg["gastos_pub"] or 0
    margen_bruto     = ingresos_mes - costo_chapa_mes - costo_mo_mes
    margen_neto      = margen_bruto - comisiones_mes - gastos_pub_mes

    df_pend_saldo = df_all[df_all["estado"] == "Pendiente saldo"]
    monto_pend    = df_pend_saldo["monto_total"].sum() - df_pend_saldo["monto_anticipo"].sum()

    st.markdown(f"**Mes actual:** {MESES_ES[mes_actual-1]} {anio_actual}")
    st.markdown('<hr class="thin">', unsafe_allow_html=True)

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        st.markdown(f"""<div class="kpi-card green">
            <div class="kpi-label">Ingresos del mes (IPV)</div>
            <div class="kpi-value">{fmt_ars(ingresos_mes)}</div>
            <div class="kpi-sub">{len(df_mes)} OTs cerradas</div>
        </div>""", unsafe_allow_html=True)
    with c2:
        st.markdown(f"""<div class="kpi-card blue">
            <div class="kpi-label">Órdenes abiertas</div>
            <div class="kpi-value">{ordenes_abiertas}</div>
            <div class="kpi-sub">{len(df_all[df_all['estado']=='Pendiente anticipo'])} sin anticipo</div>
        </div>""", unsafe_allow_html=True)
    with c3:
        st.markdown(f"""<div class="kpi-card orange">
            <div class="kpi-label">Kg procesados (mes)</div>
            <div class="kpi-value">{kg_mes:,.1f}</div>
            <div class="kpi-sub">kg de chapa</div>
        </div>""", unsafe_allow_html=True)
    with c4:
        color_mn = "green" if margen_neto >= 0 else "red"
        st.markdown(f"""<div class="kpi-card {color_mn}">
            <div class="kpi-label">Margen neto estimado</div>
            <div class="kpi-value">{fmt_ars(margen_neto)}</div>
            <div class="kpi-sub">{(margen_neto/ingresos_mes*100 if ingresos_mes else 0):.1f}% sobre IPV</div>
        </div>""", unsafe_allow_html=True)

    c5, c6, c7, c8 = st.columns(4)
    with c5:
        st.markdown(f"""<div class="kpi-card purple">
            <div class="kpi-label">Margen bruto mes</div>
            <div class="kpi-value">{fmt_ars(margen_bruto)}</div>
            <div class="kpi-sub">IPV − chapa − MO</div>
        </div>""", unsafe_allow_html=True)
    with c6:
        st.markdown(f"""<div class="kpi-card teal">
            <div class="kpi-label">Saldo pendiente</div>
            <div class="kpi-value">{fmt_ars(monto_pend)}</div>
            <div class="kpi-sub">{len(df_pend_saldo)} OTs en espera</div>
        </div>""", unsafe_allow_html=True)
    with c7:
        total_ots = len(df_all)
        st.markdown(f"""<div class="kpi-card blue">
            <div class="kpi-label">Total OTs históricas</div>
            <div class="kpi-value">{total_ots}</div>
            <div class="kpi-sub">Activas + cerradas</div>
        </div>""", unsafe_allow_html=True)
    with c8:
        prom_orden = df_mes["monto_total"].mean() if not df_mes.empty else 0
        st.markdown(f"""<div class="kpi-card orange">
            <div class="kpi-label">Promedio por orden</div>
            <div class="kpi-value">{fmt_ars(prom_orden)}</div>
            <div class="kpi-sub">OTs cerradas en el mes</div>
        </div>""", unsafe_allow_html=True)

    st.markdown('<hr class="thin">', unsafe_allow_html=True)
    col_g1, col_g2 = st.columns(2)

    with col_g1:
        st.markdown("**Estado de OTs activas**")
        if not df_activas.empty:
            counts = df_activas["estado"].value_counts().reset_index()
            counts.columns = ["Estado", "Cantidad"]
            fig = px.pie(counts, names="Estado", values="Cantidad",
                         color_discrete_sequence=["#f97316","#3b82f6","#22c55e"], hole=0.45)
            fig.update_layout(margin=dict(t=20,b=20,l=20,r=20), height=260)
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("No hay OTs activas.")

    with col_g2:
        st.markdown("**Últimas 10 OTs**")
        if not df_all.empty:
            df_show = df_all[["id_ot","cliente","estado","monto_total","fecha_pedido"]].head(10).copy()
            df_show["monto_total"] = df_show["monto_total"].apply(fmt_ars)
            df_show["fecha_pedido"] = df_all["fecha_pedido"].head(10).dt.strftime("%d/%m/%Y")
            df_show.columns = ["OT","Cliente","Estado","Monto","Fecha"]
            st.dataframe(df_show, use_container_width=True, hide_index=True, height=280)
        else:
            st.info("Sin órdenes registradas aún.")

    hoy_ts  = pd.Timestamp(hoy)
    df_venc = df_all[
        (df_all["estado"] != "Cerrada") &
        df_all["fecha_entrega"].notna() &
        (df_all["fecha_entrega"] < hoy_ts)
    ]
    if not df_venc.empty:
        st.warning(f"⚠️ {len(df_venc)} orden(es) con fecha de entrega vencida sin cerrar.")

    # Botón de descarga del Excel
    st.markdown('<hr class="thin">', unsafe_allow_html=True)
    if os.path.exists(EXCEL_PATH):
        with open(EXCEL_PATH, "rb") as f:
            st.download_button(
                "⬇️ Descargar base de datos Excel",
                f.read(),
                file_name=f"plasmart_ordenes_{date.today()}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )


# ─────────────────────────────────────────────────────────────────────────────
# ══════════════════════════ MÓDULO ÓRDENES ═══════════════════════════════════
# ─────────────────────────────────────────────────────────────────────────────
elif pagina == "📋 Órdenes de Trabajo":
    st.markdown('<div class="section-title">📋 Órdenes de Trabajo</div>', unsafe_allow_html=True)

    tab_nueva, tab_lista, tab_editar, tab_arch = st.tabs(
        ["➕ Nueva OT", "📄 Listado", "✏️ Editar OT", "🗄️ Archivadas"]
    )

    cfg_g          = load_config_global()
    vendedores_cfg = cfg_g.get("vendedores", "Vendedor").split(",")
    vendedores_cfg = [v.strip() for v in vendedores_cfg if v.strip()]
    iva_default    = float(cfg_g.get("iva_default", 21))

    # Cargar lista de clientes
    df_clientes    = load_clientes()
    clientes_lista = df_clientes["nombre"].tolist() if not df_clientes.empty else []

    # ── NUEVA OT ─────────────────────────────────────────────────────────────
    with tab_nueva:
        st.markdown("**Completá los datos de la nueva orden**")

        # Selector de cliente
        st.markdown("##### Cliente")
        modo_cliente = st.radio(
            "Modo cliente",
            ["Seleccionar cliente existente", "Nuevo cliente"],
            horizontal=True,
            label_visibility="collapsed"
        )

        cliente_nombre = ""
        cliente_cuit   = ""

        if modo_cliente == "Seleccionar cliente existente":
            if clientes_lista:
                cliente_sel = st.selectbox("Cliente", clientes_lista)
                cliente_nombre = cliente_sel
                # Traer CUIT guardado
                row_c = df_clientes[df_clientes["nombre"] == cliente_sel]
                if not row_c.empty and row_c.iloc[0]["cuit"]:
                    cliente_cuit = row_c.iloc[0]["cuit"]
                    st.caption(f"CUIT guardado: {cliente_cuit}")
            else:
                st.info("No hay clientes guardados aún. Usá la opción 'Nuevo cliente'.")
                modo_cliente = "Nuevo cliente"

        if modo_cliente == "Nuevo cliente":
            col_nc1, col_nc2 = st.columns(2)
            with col_nc1:
                cliente_nombre = st.text_input("Nombre del cliente *", placeholder="Ej: Metalúrgica López")
            with col_nc2:
                guardar_cliente = st.checkbox("Guardar en lista de clientes", value=True)

        st.markdown('<hr class="thin">', unsafe_allow_html=True)

        # Formulario principal
        with st.form("form_nueva_ot", clear_on_submit=True):
            st.markdown("##### Datos de la orden")
            n1, n2 = st.columns(2)

            with n1:
                fecha_pedido  = st.date_input("Fecha de pedido *", value=date.today())
                tipo_trabajo  = st.selectbox("Tipo de trabajo *", TIPOS_TRABAJO)
                diseno        = st.selectbox("Diseño", DISENOS)
                canal         = st.selectbox("Canal de venta", CANALES)
                vendedor      = st.selectbox("Vendedor", vendedores_cfg) if vendedores_cfg else st.text_input("Vendedor")
                kg_chapa      = st.number_input("Kg de chapa *", min_value=0.0, step=0.5, format="%.2f")
                fecha_entrega = st.date_input("Fecha de entrega estimada", value=None)

            with n2:
                monto_total  = st.number_input("Monto total (IPV) *", min_value=0.0, step=100.0, format="%.2f")
                pct_anticipo = st.slider("% Anticipo", 0, 100, 50)
                monto_anticipo_calc = monto_total * pct_anticipo / 100
                st.markdown(f"<div class='info-box'>💰 Anticipo calculado: <b>{fmt_ars(monto_anticipo_calc)}</b></div>", unsafe_allow_html=True)
                fecha_anticipo = st.date_input("Fecha de pago del anticipo", value=None)
                fecha_saldo    = st.date_input("Fecha de pago del saldo (si ya se cobró)", value=None)

                # Facturación
                st.markdown("##### Facturación")
                con_factura = st.toggle("Con factura", value=False)

                cuit_form   = ""
                monto_iva   = 0.0
                total_cf    = monto_total
                iva_pct_val = iva_default

                if con_factura:
                    cuit_form   = st.text_input("CUIT del cliente *", value=cliente_cuit, placeholder="20-12345678-9")
                    iva_pct_val = st.number_input("IVA %", value=iva_default, min_value=0.0, max_value=100.0, step=0.5)
                    monto_iva   = monto_total * iva_pct_val / 100
                    total_cf    = monto_total + monto_iva
                    st.markdown(f"""<div class='factura-box'>
                        🧾 IVA ({iva_pct_val:.0f}%): <b>{fmt_ars(monto_iva)}</b><br>
                        Total con IVA: <b>{fmt_ars(total_cf)}</b>
                    </div>""", unsafe_allow_html=True)
                else:
                    st.markdown("<div class='sin-factura-box'>📄 Venta sin factura</div>", unsafe_allow_html=True)

                notas = st.text_area("Notas (opcional)", height=80)

            submitted = st.form_submit_button("✅ Crear Orden de Trabajo", use_container_width=True, type="primary")

        if submitted:
            errores = []
            if not cliente_nombre.strip():
                errores.append("El nombre del cliente es obligatorio.")
            if monto_total <= 0:
                errores.append("El monto total debe ser mayor a 0.")
            if kg_chapa <= 0:
                errores.append("Los kg de chapa deben ser mayores a 0.")
            if con_factura and not cuit_form.strip():
                errores.append("El CUIT es obligatorio cuando la venta es con factura.")

            if errores:
                for e in errores:
                    st.error(e)
            else:
                ot_id  = next_ot_id()
                estado = calc_estado(
                    fecha_anticipo.isoformat() if fecha_anticipo else None,
                    fecha_saldo.isoformat()    if fecha_saldo    else None
                )

                # Guardar OT
                conn = get_conn()
                conn.execute("""
                    INSERT INTO ordenes
                    (id_ot,cliente,cuit,fecha_pedido,tipo_trabajo,diseno,
                     monto_total,con_factura,iva_pct,monto_iva,monto_total_cf,
                     pct_anticipo,monto_anticipo,fecha_anticipo,
                     kg_chapa,fecha_entrega,fecha_saldo,canal,vendedor,estado,notas)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """, (
                    ot_id,
                    cliente_nombre.strip(),
                    cuit_form.strip() if con_factura else None,
                    fecha_pedido.isoformat(),
                    tipo_trabajo, diseno,
                    monto_total,
                    1 if con_factura else 0,
                    iva_pct_val if con_factura else 0,
                    monto_iva, total_cf,
                    pct_anticipo, monto_anticipo_calc,
                    fecha_anticipo.isoformat() if fecha_anticipo else None,
                    kg_chapa,
                    fecha_entrega.isoformat() if fecha_entrega else None,
                    fecha_saldo.isoformat()    if fecha_saldo    else None,
                    canal, vendedor, estado,
                    notas.strip() or None
                ))

                # Guardar cliente nuevo si corresponde
                if modo_cliente == "Nuevo cliente" and cliente_nombre.strip():
                    try:
                        conn.execute(
                            "INSERT OR IGNORE INTO clientes (nombre, cuit) VALUES (?,?)",
                            (cliente_nombre.strip(), cuit_form.strip() if con_factura else None)
                        )
                    except Exception:
                        pass

                conn.commit()
                conn.close()

                # Actualizar Excel automáticamente
                actualizar_excel()

                st.success(f"✅ Orden **{ot_id}** creada para **{cliente_nombre.strip()}**. Excel actualizado.")
                st.balloons()

    # ── LISTADO ──────────────────────────────────────────────────────────────
    with tab_lista:
        df = load_ordenes()

        with st.expander("🔍 Filtros", expanded=False):
            fc1, fc2, fc3, fc4, fc5 = st.columns(5)
            with fc1:
                f_estado = st.multiselect("Estado", ESTADOS, default=ESTADOS)
            with fc2:
                f_canal  = st.multiselect("Canal", CANALES, default=CANALES)
            with fc3:
                vendedores_df = sorted(df["vendedor"].dropna().unique().tolist()) if not df.empty else []
                f_vend = st.multiselect("Vendedor", vendedores_df, default=vendedores_df)
            with fc4:
                f_desde = st.date_input("Desde", value=None, key="lista_desde")
            with fc5:
                f_hasta = st.date_input("Hasta", value=None, key="lista_hasta")
            f_cliente  = st.text_input("Buscar cliente", "")
            f_factura  = st.selectbox("Facturación", ["Todas", "Con factura", "Sin factura"])
            f_trabajo  = st.multiselect("Tipo de trabajo", TIPOS_TRABAJO, default=TIPOS_TRABAJO)

        df_f = df.copy()
        if f_estado:
            df_f = df_f[df_f["estado"].isin(f_estado)]
        if f_canal:
            df_f = df_f[df_f["canal"].isin(f_canal)]
        if f_vend:
            df_f = df_f[df_f["vendedor"].isin(f_vend)]
        if f_desde:
            df_f = df_f[df_f["fecha_pedido"] >= pd.Timestamp(f_desde)]
        if f_hasta:
            df_f = df_f[df_f["fecha_pedido"] <= pd.Timestamp(f_hasta)]
        if f_cliente:
            df_f = df_f[df_f["cliente"].str.contains(f_cliente, case=False, na=False)]
        if f_factura == "Con factura":
            df_f = df_f[df_f["con_factura"] == 1]
        elif f_factura == "Sin factura":
            df_f = df_f[df_f["con_factura"] == 0]
        if f_trabajo:
            df_f = df_f[df_f["tipo_trabajo"].isin(f_trabajo)]

        st.markdown(f"**{len(df_f)} órdenes encontradas**")

        if not df_f.empty:
            cols_show = ["id_ot","cliente","tipo_trabajo","estado","monto_total",
                         "con_factura","monto_iva","kg_chapa","canal","vendedor","fecha_pedido"]
            cols_pres = [c for c in cols_show if c in df_f.columns]
            df_show   = df_f[cols_pres].copy()
            df_show["monto_total"] = df_show["monto_total"].apply(fmt_ars)
            df_show["monto_iva"]   = df_show["monto_iva"].apply(fmt_ars)   if "monto_iva"   in df_show.columns else "—"
            df_show["kg_chapa"]    = df_show["kg_chapa"].apply(fmt_kg)
            df_show["con_factura"] = df_show["con_factura"].apply(lambda x: "✅ Sí" if x == 1 else "—")
            df_show["fecha_pedido"]= df_f["fecha_pedido"].dt.strftime("%d/%m/%Y")
            df_show.columns        = ["OT","Cliente","Trabajo","Estado","Monto","Factura",
                                       "IVA","Kg","Canal","Vendedor","Fecha"]
            st.dataframe(df_show, use_container_width=True, hide_index=True, height=420)

            col_dl1, col_dl2 = st.columns(2)
            with col_dl1:
                st.download_button(
                    "⬇️ Exportar selección a Excel",
                    to_excel_download(df_f),
                    file_name=f"plasmart_filtrado_{date.today()}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            with col_dl2:
                if os.path.exists(EXCEL_PATH):
                    with open(EXCEL_PATH, "rb") as f_xls:
                        st.download_button(
                            "⬇️ Descargar base completa Excel",
                            f_xls.read(),
                            file_name=f"plasmart_ordenes_{date.today()}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
        else:
            st.info("No se encontraron órdenes con los filtros seleccionados.")

    # ── EDITAR OT ─────────────────────────────────────────────────────────────
    with tab_editar:
        df_edit = load_ordenes()
        if df_edit.empty:
            st.info("No hay órdenes para editar.")
        else:
            ot_sel = st.selectbox("Seleccioná la OT a editar", df_edit["id_ot"].tolist())
            row    = df_edit[df_edit["id_ot"] == ot_sel].iloc[0]

            with st.form(f"form_editar_{ot_sel}"):
                e1, e2 = st.columns(2)
                with e1:
                    e_cliente  = st.text_input("Cliente", value=row["cliente"])
                    e_cuit     = st.text_input("CUIT", value=row["cuit"] or "")
                    e_fecha_p  = st.date_input("Fecha de pedido",
                        value=row["fecha_pedido"].date() if pd.notna(row["fecha_pedido"]) else date.today())
                    tt_idx     = TIPOS_TRABAJO.index(row["tipo_trabajo"]) if row.get("tipo_trabajo") in TIPOS_TRABAJO else 0
                    e_tipo     = st.selectbox("Tipo de trabajo", TIPOS_TRABAJO, index=tt_idx)
                    dis_idx    = DISENOS.index(row["diseno"]) if row.get("diseno") in DISENOS else 0
                    e_diseno   = st.selectbox("Diseño", DISENOS, index=dis_idx)
                    e_kg       = st.number_input("Kg chapa", value=float(row["kg_chapa"]), min_value=0.0, step=0.5)
                    val_fe     = row["fecha_entrega"].date() if pd.notna(row["fecha_entrega"]) else None
                    e_fecha_e  = st.date_input("Fecha entrega estimada", value=val_fe)

                with e2:
                    e_monto    = st.number_input("Monto total (IPV)", value=float(row["monto_total"]), min_value=0.0, step=100.0)
                    e_pct      = st.slider("% Anticipo", 0, 100, int(row["pct_anticipo"]))
                    e_monto_a  = e_monto * e_pct / 100
                    st.markdown(f"<div class='info-box'>Anticipo calculado: <b>{fmt_ars(e_monto_a)}</b></div>", unsafe_allow_html=True)
                    val_fa     = row["fecha_anticipo"].date() if pd.notna(row["fecha_anticipo"]) else None
                    e_fecha_a  = st.date_input("Fecha pago anticipo", value=val_fa)
                    val_fs     = row["fecha_saldo"].date() if pd.notna(row["fecha_saldo"]) else None
                    e_fecha_s  = st.date_input("Fecha pago saldo", value=val_fs)
                    canal_idx  = CANALES.index(row["canal"]) if row["canal"] in CANALES else 0
                    e_canal    = st.selectbox("Canal", CANALES, index=canal_idx)
                    vend_idx   = vendedores_cfg.index(row["vendedor"]) if row["vendedor"] in vendedores_cfg else 0
                    e_vendedor = st.selectbox("Vendedor", vendedores_cfg, index=vend_idx) if vendedores_cfg else st.text_input("Vendedor", value=row["vendedor"] or "")

                    e_con_fac  = st.toggle("Con factura", value=bool(row.get("con_factura", 0)))
                    e_iva_monto = 0.0
                    e_total_cf  = e_monto
                    e_iva_pct   = float(row.get("iva_pct") or iva_default)
                    if e_con_fac:
                        e_iva_pct   = st.number_input("IVA %", value=e_iva_pct, min_value=0.0, max_value=100.0, step=0.5)
                        e_iva_monto = e_monto * e_iva_pct / 100
                        e_total_cf  = e_monto + e_iva_monto
                        st.markdown(f"<div class='factura-box'>IVA: <b>{fmt_ars(e_iva_monto)}</b> | Total: <b>{fmt_ars(e_total_cf)}</b></div>", unsafe_allow_html=True)

                    e_notas = st.text_area("Notas", value=row["notas"] or "", height=80)

                col_b1, col_b2 = st.columns([3,1])
                with col_b1:
                    save_btn = st.form_submit_button("💾 Guardar cambios", type="primary", use_container_width=True)
                with col_b2:
                    del_btn  = st.form_submit_button("🗄️ Archivar", use_container_width=True)

            if save_btn:
                nuevo_estado = calc_estado(
                    e_fecha_a.isoformat() if e_fecha_a else None,
                    e_fecha_s.isoformat() if e_fecha_s else None
                )
                conn = get_conn()
                conn.execute("""
                    UPDATE ordenes SET
                        cliente=?, cuit=?, fecha_pedido=?, tipo_trabajo=?, diseno=?,
                        monto_total=?, con_factura=?, iva_pct=?, monto_iva=?, monto_total_cf=?,
                        pct_anticipo=?, monto_anticipo=?, fecha_anticipo=?,
                        kg_chapa=?, fecha_entrega=?, fecha_saldo=?,
                        canal=?, vendedor=?, estado=?, notas=?
                    WHERE id_ot=?
                """, (
                    e_cliente.strip(),
                    e_cuit.strip() if e_con_fac else None,
                    e_fecha_p.isoformat(),
                    e_tipo, e_diseno,
                    e_monto,
                    1 if e_con_fac else 0,
                    e_iva_pct if e_con_fac else 0,
                    e_iva_monto, e_total_cf,
                    e_pct, e_monto_a,
                    e_fecha_a.isoformat() if e_fecha_a else None,
                    e_kg,
                    e_fecha_e.isoformat() if e_fecha_e else None,
                    e_fecha_s.isoformat() if e_fecha_s else None,
                    e_canal, e_vendedor, nuevo_estado,
                    e_notas.strip() or None,
                    ot_sel
                ))
                conn.commit()
                conn.close()
                actualizar_excel()
                st.success(f"✅ OT {ot_sel} actualizada. Estado: **{nuevo_estado}**. Excel actualizado.")
                st.rerun()

            if del_btn:
                conn = get_conn()
                conn.execute("UPDATE ordenes SET archivada=1 WHERE id_ot=?", (ot_sel,))
                conn.commit()
                conn.close()
                actualizar_excel()
                st.warning(f"🗄️ OT {ot_sel} archivada.")
                st.rerun()

    # ── ARCHIVADAS ────────────────────────────────────────────────────────────
    with tab_arch:
        df_arch = load_ordenes(archivadas=True)
        if df_arch.empty:
            st.info("No hay órdenes archivadas.")
        else:
            st.markdown(f"**{len(df_arch)} órdenes archivadas**")
            df_arch_show = df_arch[["id_ot","cliente","estado","monto_total","fecha_pedido"]].copy()
            df_arch_show["monto_total"]  = df_arch_show["monto_total"].apply(fmt_ars)
            df_arch_show["fecha_pedido"] = df_arch["fecha_pedido"].dt.strftime("%d/%m/%Y")
            df_arch_show.columns = ["OT","Cliente","Estado","Monto","Fecha"]
            st.dataframe(df_arch_show, use_container_width=True, hide_index=True)

            ot_rest = st.selectbox("Restaurar OT", df_arch["id_ot"].tolist(), key="rest_arch")
            if st.button("↩️ Restaurar"):
                conn = get_conn()
                conn.execute("UPDATE ordenes SET archivada=0 WHERE id_ot=?", (ot_rest,))
                conn.commit()
                conn.close()
                actualizar_excel()
                st.success(f"✅ OT {ot_rest} restaurada.")
                st.rerun()


# ─────────────────────────────────────────────────────────────────────────────
# ════════════════════════════ CLIENTES ═══════════════════════════════════════
# ─────────────────────────────────────────────────────────────────────────────
elif pagina == "👥 Clientes":
    st.markdown('<div class="section-title">👥 Clientes</div>', unsafe_allow_html=True)

    tab_lista_c, tab_nuevo_c = st.tabs(["📄 Lista de clientes", "➕ Nuevo cliente"])

    with tab_lista_c:
        df_c = load_clientes()
        if df_c.empty:
            st.info("No hay clientes guardados aún.")
        else:
            st.markdown(f"**{len(df_c)} clientes registrados**")
            # Agregar historial de OTs por cliente
            df_ots = load_ordenes()
            if not df_ots.empty:
                conteo = df_ots.groupby("cliente").agg(
                    OTs=("id_ot","count"),
                    Total=("monto_total","sum")
                ).reset_index()
                df_c = df_c.merge(conteo, left_on="nombre", right_on="cliente", how="left")
                df_c["OTs"]   = df_c["OTs"].fillna(0).astype(int)
                df_c["Total"] = df_c["Total"].fillna(0).apply(fmt_ars)
            df_show = df_c[["nombre","cuit","telefono","email","OTs","Total"]].copy() if "OTs" in df_c.columns else df_c[["nombre","cuit","telefono","email"]]
            df_show.columns = ["Nombre","CUIT","Teléfono","Email","OTs","Facturado"] if "OTs" in df_c.columns else ["Nombre","CUIT","Teléfono","Email"]
            st.dataframe(df_show, use_container_width=True, hide_index=True)

            st.download_button(
                "⬇️ Exportar clientes a Excel",
                to_excel_download(df_c),
                file_name=f"plasmart_clientes_{date.today()}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    with tab_nuevo_c:
        with st.form("form_nuevo_cliente"):
            cc1, cc2 = st.columns(2)
            with cc1:
                c_nombre   = st.text_input("Nombre / Razón social *")
                c_cuit     = st.text_input("CUIT", placeholder="20-12345678-9")
            with cc2:
                c_telefono = st.text_input("Teléfono")
                c_email    = st.text_input("Email")
            if st.form_submit_button("✅ Guardar cliente", type="primary"):
                if not c_nombre.strip():
                    st.error("El nombre es obligatorio.")
                else:
                    conn = get_conn()
                    try:
                        conn.execute(
                            "INSERT INTO clientes (nombre,cuit,telefono,email) VALUES (?,?,?,?)",
                            (c_nombre.strip(), c_cuit.strip() or None,
                             c_telefono.strip() or None, c_email.strip() or None)
                        )
                        conn.commit()
                        st.success(f"✅ Cliente **{c_nombre.strip()}** guardado.")
                        st.rerun()
                    except sqlite3.IntegrityError:
                        st.error("Ya existe un cliente con ese nombre.")
                    finally:
                        conn.close()


# ─────────────────────────────────────────────────────────────────────────────
# ════════════════════════ ANÁLISIS MENSUAL ═══════════════════════════════════
# ─────────────────────────────────────────────────────────────────────────────
elif pagina == "📊 Análisis Mensual":
    st.markdown('<div class="section-title">📊 Análisis Mensual</div>', unsafe_allow_html=True)

    hoy = date.today()
    col_sel1, col_sel2 = st.columns([1,1])
    with col_sel1:
        mes_sel  = st.selectbox("Mes", range(1,13), index=hoy.month-1, format_func=lambda x: MESES_ES[x-1])
    with col_sel2:
        anio_sel = st.selectbox("Año", range(2023, hoy.year+2), index=hoy.year-2023)

    st.markdown('<hr class="thin">', unsafe_allow_html=True)

    cfg       = load_config_mensual(anio_sel, mes_sel)
    cfg_g     = load_config_global()
    costo_kg_default = float(cfg_g.get("costo_chapa_kg_default", 800))

    with st.expander("⚙️ Configuración de costos del mes", expanded=False):
        with st.form("form_config_mensual"):
            cm1, cm2, cm3 = st.columns(3)
            with cm1:
                c_chapa = st.number_input("Costo chapa ($/kg)", value=float(cfg["costo_chapa_kg"] or costo_kg_default), min_value=0.0, step=10.0)
            with cm2:
                c_mo    = st.number_input("Costo Mano de Obra ($)", value=float(cfg["costo_mo"] or 0), min_value=0.0, step=1000.0)
            with cm3:
                c_pub   = st.number_input("Gastos publicitarios ($)", value=float(cfg["gastos_pub"] or 0), min_value=0.0, step=1000.0)
            if st.form_submit_button("💾 Guardar", type="primary"):
                save_config_mensual(anio_sel, mes_sel, c_chapa, c_mo, c_pub)
                st.success("Configuración guardada.")
                st.rerun()

    df_all = load_ordenes()
    df_mes = df_all[
        (df_all["estado"] == "Cerrada") &
        (df_all["fecha_saldo"].dt.month == mes_sel) &
        (df_all["fecha_saldo"].dt.year  == anio_sel)
    ].copy()

    cfg_fresh  = load_config_mensual(anio_sel, mes_sel)
    costo_kg   = cfg_fresh["costo_chapa_kg"] or costo_kg_default
    ipv_total  = df_mes["monto_total"].sum()
    kg_total   = df_mes["kg_chapa"].sum()
    costo_chapa = kg_total * costo_kg
    costo_mo   = cfg_fresh["costo_mo"] or 0
    gastos_pub = cfg_fresh["gastos_pub"] or 0
    comisiones = ipv_total * 0.01
    margen_bruto = ipv_total - costo_chapa - costo_mo
    margen_neto  = margen_bruto - comisiones - gastos_pub

    st.markdown(f"**{MESES_ES[mes_sel-1]} {anio_sel}** — {len(df_mes)} órdenes cerradas")

    r1c1,r1c2,r1c3,r1c4 = st.columns(4)
    with r1c1:
        st.markdown(f"""<div class="kpi-card green"><div class="kpi-label">Ingresos totales (IPV)</div>
            <div class="kpi-value">{fmt_ars(ipv_total)}</div><div class="kpi-sub">{len(df_mes)} OTs cerradas</div></div>""", unsafe_allow_html=True)
    with r1c2:
        st.markdown(f"""<div class="kpi-card orange"><div class="kpi-label">Total Kg de chapa</div>
            <div class="kpi-value">{kg_total:,.1f}</div><div class="kpi-sub">kg procesados</div></div>""", unsafe_allow_html=True)
    with r1c3:
        st.markdown(f"""<div class="kpi-card red"><div class="kpi-label">Costo de chapa</div>
            <div class="kpi-value">{fmt_ars(costo_chapa)}</div><div class="kpi-sub">@ {fmt_ars(costo_kg)}/kg</div></div>""", unsafe_allow_html=True)
    with r1c4:
        st.markdown(f"""<div class="kpi-card red"><div class="kpi-label">Costo Mano de Obra</div>
            <div class="kpi-value">{fmt_ars(costo_mo)}</div><div class="kpi-sub">según configuración</div></div>""", unsafe_allow_html=True)

    r2c1,r2c2,r2c3,r2c4 = st.columns(4)
    with r2c1:
        st.markdown(f"""<div class="kpi-card purple"><div class="kpi-label">Margen bruto</div>
            <div class="kpi-value">{fmt_ars(margen_bruto)}</div><div class="kpi-sub">IPV − chapa − MO</div></div>""", unsafe_allow_html=True)
    with r2c2:
        st.markdown(f"""<div class="kpi-card orange"><div class="kpi-label">Comisiones (1%)</div>
            <div class="kpi-value">{fmt_ars(comisiones)}</div><div class="kpi-sub">1% sobre IPV</div></div>""", unsafe_allow_html=True)
    with r2c3:
        st.markdown(f"""<div class="kpi-card blue"><div class="kpi-label">Gastos publicitarios</div>
            <div class="kpi-value">{fmt_ars(gastos_pub)}</div><div class="kpi-sub">según configuración</div></div>""", unsafe_allow_html=True)
    with r2c4:
        color_mn = "green" if margen_neto >= 0 else "red"
        st.markdown(f"""<div class="kpi-card {color_mn}"><div class="kpi-label">Margen neto</div>
            <div class="kpi-value">{fmt_ars(margen_neto)}</div>
            <div class="kpi-sub">{(margen_neto/ipv_total*100 if ipv_total else 0):.1f}% sobre IPV</div></div>""", unsafe_allow_html=True)

    if not df_mes.empty:
        st.markdown('<hr class="thin">', unsafe_allow_html=True)
        st.markdown("**Detalle de OTs del mes**")
        cols_m = ["id_ot","cliente","tipo_trabajo","monto_total","con_factura",
                  "monto_iva","kg_chapa","canal","vendedor","fecha_saldo","notas"]
        cols_m = [c for c in cols_m if c in df_mes.columns]
        df_mes_show = df_mes[cols_m].copy()
        df_mes_show["monto_total"] = df_mes_show["monto_total"].apply(fmt_ars)
        df_mes_show["monto_iva"]   = df_mes_show["monto_iva"].apply(fmt_ars) if "monto_iva" in df_mes_show.columns else "—"
        df_mes_show["kg_chapa"]    = df_mes_show["kg_chapa"].apply(fmt_kg)
        df_mes_show["con_factura"] = df_mes_show["con_factura"].apply(lambda x: "✅" if x == 1 else "—")
        df_mes_show["fecha_saldo"] = df_mes["fecha_saldo"].dt.strftime("%d/%m/%Y")
        st.dataframe(df_mes_show, use_container_width=True, hide_index=True)

        st.download_button(
            "⬇️ Exportar mes a Excel",
            to_excel_download(df_mes),
            file_name=f"plasmart_{MESES_ES[mes_sel-1]}_{anio_sel}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    else:
        st.info(f"No hay órdenes cerradas en {MESES_ES[mes_sel-1]} {anio_sel}.")


# ─────────────────────────────────────────────────────────────────────────────
# ════════════════════════ ANÁLISIS ANUAL ═════════════════════════════════════
# ─────────────────────────────────────────────────────────────────────────────
elif pagina == "📈 Análisis Anual":
    st.markdown('<div class="section-title">📈 Análisis Anual</div>', unsafe_allow_html=True)

    hoy      = date.today()
    anio_sel = st.selectbox("Año", range(2023, hoy.year+2), index=hoy.year-2023)

    df_all   = load_ordenes()
    cfg_g    = load_config_global()
    costo_kg_default = float(cfg_g.get("costo_chapa_kg_default", 800))

    df_anio  = df_all[
        (df_all["estado"] == "Cerrada") &
        (df_all["fecha_saldo"].dt.year == anio_sel)
    ].copy()
    df_anio["mes_num"] = df_anio["fecha_saldo"].dt.month
    df_anio["mes_nom"] = df_anio["mes_num"].apply(lambda x: MESES_ES[x-1])

    registros = []
    for mes in range(1,13):
        dm   = df_anio[df_anio["mes_num"] == mes]
        cfg_m = load_config_mensual(anio_sel, mes)
        ck   = cfg_m["costo_chapa_kg"] or costo_kg_default
        ipv  = dm["monto_total"].sum()
        kg   = dm["kg_chapa"].sum()
        cc   = kg * ck
        mo   = cfg_m["costo_mo"] or 0
        gp   = cfg_m["gastos_pub"] or 0
        com  = ipv * 0.01
        mb   = ipv - cc - mo
        mn   = mb - com - gp
        registros.append({
            "Mes": MESES_ES[mes-1], "Mes_num": mes,
            "IPV": ipv, "Kg": kg, "Costo_chapa": cc,
            "Costo_MO": mo, "Comisiones": com, "Gastos_pub": gp,
            "Margen_bruto": mb, "Margen_neto": mn,
            "Num_OTs": len(dm),
            "Prom_orden": ipv/len(dm) if len(dm) > 0 else 0
        })
    df_resumen = pd.DataFrame(registros)

    ipv_anual  = df_resumen["IPV"].sum()
    kg_anual   = df_resumen["Kg"].sum()
    mn_anual   = df_resumen["Margen_neto"].sum()
    mb_anual   = df_resumen["Margen_bruto"].sum()
    total_ots  = df_resumen["Num_OTs"].sum()
    prom_orden = ipv_anual / total_ots if total_ots > 0 else 0

    df_ant     = df_all[
        (df_all["estado"] == "Cerrada") &
        (df_all["fecha_saldo"].dt.year == anio_sel - 1)
    ]
    ipv_ant  = df_ant["monto_total"].sum() if not df_ant.empty else 0
    pct_crec = ((ipv_anual - ipv_ant) / ipv_ant * 100) if ipv_ant > 0 else None

    ka1,ka2,ka3,ka4,ka5 = st.columns(5)
    with ka1:
        st.markdown(f"""<div class="kpi-card green"><div class="kpi-label">Facturación anual</div>
            <div class="kpi-value">{fmt_ars(ipv_anual)}</div><div class="kpi-sub">{total_ots} OTs cerradas</div></div>""", unsafe_allow_html=True)
    with ka2:
        st.markdown(f"""<div class="kpi-card orange"><div class="kpi-label">Kg totales</div>
            <div class="kpi-value">{kg_anual:,.0f}</div><div class="kpi-sub">kg procesados</div></div>""", unsafe_allow_html=True)
    with ka3:
        st.markdown(f"""<div class="kpi-card purple"><div class="kpi-label">Margen bruto anual</div>
            <div class="kpi-value">{fmt_ars(mb_anual)}</div>
            <div class="kpi-sub">{(mb_anual/ipv_anual*100 if ipv_anual else 0):.1f}%</div></div>""", unsafe_allow_html=True)
    with ka4:
        color_mn = "green" if mn_anual >= 0 else "red"
        st.markdown(f"""<div class="kpi-card {color_mn}"><div class="kpi-label">Margen neto anual</div>
            <div class="kpi-value">{fmt_ars(mn_anual)}</div>
            <div class="kpi-sub">{(mn_anual/ipv_anual*100 if ipv_anual else 0):.1f}%</div></div>""", unsafe_allow_html=True)
    with ka5:
        crec_txt = f"{pct_crec:+.1f}%" if pct_crec is not None else "N/D"
        crec_col = "green" if (pct_crec or 0) >= 0 else "red"
        st.markdown(f"""<div class="kpi-card {crec_col}"><div class="kpi-label">Crecimiento vs {anio_sel-1}</div>
            <div class="kpi-value">{crec_txt}</div><div class="kpi-sub">Facturación IPV</div></div>""", unsafe_allow_html=True)

    st.markdown('<hr class="thin">', unsafe_allow_html=True)

    if not df_anio.empty:
        g1, g2 = st.columns(2)
        with g1:
            st.markdown("**Evolución mensual — IPV y márgenes**")
            df_plot = df_resumen[df_resumen["Num_OTs"] > 0]
            fig1    = go.Figure()
            fig1.add_trace(go.Bar(name="IPV", x=df_plot["Mes"], y=df_plot["IPV"], marker_color="#3b82f6", opacity=0.7))
            fig1.add_trace(go.Scatter(name="Mg. Bruto", x=df_plot["Mes"], y=df_plot["Margen_bruto"], mode="lines+markers", line=dict(color="#f97316", width=2)))
            fig1.add_trace(go.Scatter(name="Mg. Neto",  x=df_plot["Mes"], y=df_plot["Margen_neto"],  mode="lines+markers", line=dict(color="#22c55e", width=2, dash="dot")))
            fig1.update_layout(height=300, margin=dict(t=10,b=20,l=20,r=20), plot_bgcolor="white", paper_bgcolor="white",
                               legend=dict(orientation="h",y=-0.25), yaxis_tickformat="$,.0f")
            st.plotly_chart(fig1, use_container_width=True)

        with g2:
            st.markdown("**Ventas por canal**")
            df_canal = df_anio.groupby("canal")["monto_total"].sum().reset_index()
            df_canal.columns = ["Canal","IPV"]
            fig3 = px.bar(df_canal.sort_values("IPV", ascending=True), x="IPV", y="Canal",
                          orientation="h", color="Canal",
                          color_discrete_sequence=px.colors.qualitative.Set2)
            fig3.update_layout(height=300, margin=dict(t=10,b=20,l=20,r=20), showlegend=False,
                               plot_bgcolor="white", paper_bgcolor="white", xaxis_tickformat="$,.0f")
            st.plotly_chart(fig3, use_container_width=True)

        if "tipo_trabajo" in df_anio.columns:
            g3, g4 = st.columns(2)
            with g3:
                st.markdown("**Distribución por tipo de trabajo**")
                df_tipo = df_anio.groupby("tipo_trabajo")["monto_total"].sum().reset_index()
                fig_t   = px.pie(df_tipo, names="tipo_trabajo", values="monto_total",
                                 color_discrete_sequence=px.colors.qualitative.Pastel, hole=0.4)
                fig_t.update_layout(height=260, margin=dict(t=10,b=20,l=20,r=20))
                st.plotly_chart(fig_t, use_container_width=True)
            with g4:
                st.markdown("**Kg de chapa por mes**")
                df_kg = df_resumen[df_resumen["Num_OTs"] > 0]
                fig_k = px.bar(df_kg, x="Mes", y="Kg", color_discrete_sequence=["#14b8a6"])
                fig_k.update_layout(height=260, margin=dict(t=10,b=20,l=20,r=20),
                                    plot_bgcolor="white", paper_bgcolor="white")
                st.plotly_chart(fig_k, use_container_width=True)

        st.markdown('<hr class="thin">', unsafe_allow_html=True)
        st.markdown("**Resumen mensual**")
        df_tabla = df_resumen.copy()
        for col in ["IPV","Costo_chapa","Costo_MO","Comisiones","Gastos_pub","Margen_bruto","Margen_neto","Prom_orden"]:
            df_tabla[col] = df_tabla[col].apply(fmt_ars)
        df_tabla["Kg"] = df_tabla["Kg"].apply(lambda x: f"{x:,.1f}")
        df_tabla = df_tabla.drop(columns=["Mes_num"])
        df_tabla.columns = ["Mes","IPV","Kg","Costo Chapa","Costo MO","Comisiones",
                            "G.Pub","Mg.Bruto","Mg.Neto","# OTs","Prom/OT"]
        st.dataframe(df_tabla, use_container_width=True, hide_index=True)
        st.download_button("⬇️ Exportar análisis anual", to_excel_download(df_resumen),
                           file_name=f"plasmart_anual_{anio_sel}.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    else:
        st.info(f"No hay órdenes cerradas en {anio_sel}.")


# ─────────────────────────────────────────────────────────────────────────────
# ════════════════════════ CONFIGURACIÓN ══════════════════════════════════════
# ─────────────────────────────────────────────────────────────────────────────
elif pagina == "⚙️ Configuración":
    st.markdown('<div class="section-title">⚙️ Configuración</div>', unsafe_allow_html=True)

    cfg_g = load_config_global()
    tab_gen, tab_vend = st.tabs(["General", "Vendedores"])

    with tab_gen:
        with st.form("form_config_global"):
            cg1, cg2 = st.columns(2)
            with cg1:
                costo_kg = st.number_input("Costo por kg de chapa ($ por defecto)",
                    value=float(cfg_g.get("costo_chapa_kg_default", 800)), min_value=0.0, step=10.0)
            with cg2:
                iva_def = st.number_input("IVA por defecto (%)",
                    value=float(cfg_g.get("iva_default", 21)), min_value=0.0, max_value=100.0, step=0.5)
            if st.form_submit_button("💾 Guardar", type="primary"):
                save_config_global("costo_chapa_kg_default", costo_kg)
                save_config_global("iva_default", iva_def)
                st.success("Configuración guardada.")
                st.rerun()

    with tab_vend:
        vendedores_raw  = cfg_g.get("vendedores", "")
        vendedores_lista = [v.strip() for v in vendedores_raw.split(",") if v.strip()]
        st.markdown(f"Vendedores actuales: **{', '.join(vendedores_lista)}**")
        with st.form("form_vendedores"):
            nuevos = st.text_area("Vendedores separados por coma",
                value=", ".join(vendedores_lista), height=80)
            if st.form_submit_button("💾 Guardar", type="primary"):
                lista_limpia = ", ".join([v.strip() for v in nuevos.split(",") if v.strip()])
                save_config_global("vendedores", lista_limpia)
                st.success("Lista actualizada.")
                st.rerun()

    st.markdown('<hr class="thin">', unsafe_allow_html=True)
    st.markdown("**Sistema**")
    col_i1, col_i2 = st.columns(2)
    with col_i1:
        db_size = os.path.getsize(DB_PATH) / 1024 if os.path.exists(DB_PATH) else 0
        df_all  = load_ordenes()
        st.markdown(f"""<div class="info-box">
            📦 Base de datos: <b>{DB_PATH}</b><br>
            💾 Tamaño: <b>{db_size:.1f} KB</b><br>
            📋 Total OTs activas: <b>{len(df_all)}</b>
        </div>""", unsafe_allow_html=True)
    with col_i2:
        st.markdown("""<div class="info-box">
            ⚡ <b>Plasmart Gestor de OTs v2.0</b><br>
            Córdoba, Argentina<br>
            <a href="https://plasmartcba.com" style="color:#f97316;">plasmartcba.com</a>
        </div>""", unsafe_allow_html=True)

    st.markdown('<hr class="thin">', unsafe_allow_html=True)
    st.markdown("**Exportar toda la base de datos**")
    if not df_all.empty:
        col_x1, col_x2 = st.columns(2)
        with col_x1:
            st.download_button("⬇️ Exportar todas las OTs (Excel)",
                to_excel_download(df_all),
                file_name=f"plasmart_backup_{date.today()}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary")
        with col_x2:
            if os.path.exists(EXCEL_PATH):
                with open(EXCEL_PATH, "rb") as f:
                    st.download_button("⬇️ Descargar Excel automático",
                        f.read(),
                        file_name=f"plasmart_ordenes_{date.today()}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# =============================================================================
# PLASMART - Gestor de Órdenes de Trabajo
# plasmartcba.com
# Backend: Google Sheets (gspread + oauth2client)
# =============================================================================
#
# INSTALACIÓN LOCAL:
#   pip install streamlit pandas gspread google-auth openpyxl
#
# VARIABLES DE ENTORNO requeridas (Railway → Variables):
#   GOOGLE_SHEET_ID   → ID del Google Sheet (está en la URL)
#   GOOGLE_CREDS_JSON → contenido completo del JSON de la cuenta de servicio
#
# EJECUCIÓN LOCAL:
#   streamlit run app.py
# =============================================================================

import streamlit as st
import pandas as pd
import gspread
from google.oauth2.service_account import Credentials
from datetime import date, datetime
import json
import os
import io
import time

# ---------------------------------------------------------------------------
# CONFIGURACIÓN DE PÁGINA
# ---------------------------------------------------------------------------
st.set_page_config(
    page_title="Plasmart · Órdenes de Trabajo",
    page_icon="⚙️",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ---------------------------------------------------------------------------
# ESTILOS
# ---------------------------------------------------------------------------
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Sora:wght@300;400;500;600;700&family=JetBrains+Mono:wght@400;500&display=swap');

    .stApp {
        background: #0f1117 !important;
        font-family: 'Sora', sans-serif !important;
        color: #f1f5f9 !important;
    }
    .plasmart-header {
        background: linear-gradient(135deg, #1e2130 0%, #252840 50%, #1e2130 100%);
        border: 1px solid rgba(249,115,22,0.3);
        border-radius: 16px;
        padding: 28px 36px;
        margin-bottom: 28px;
        display: flex;
        align-items: center;
        gap: 20px;
        box-shadow: 0 4px 24px rgba(0,0,0,0.4), 0 0 60px rgba(249,115,22,0.08);
        position: relative;
        overflow: hidden;
    }
    .plasmart-header::before {
        content: '';
        position: absolute;
        top: 0; left: 0; right: 0;
        height: 2px;
        background: linear-gradient(90deg, transparent, #f97316, transparent);
    }
    .plasmart-logo {
        font-size: 2.4rem;
        font-weight: 700;
        color: #f97316;
        letter-spacing: -1px;
        line-height: 1;
    }
    .plasmart-logo span { color: #f1f5f9; }
    .plasmart-subtitle {
        font-size: 0.85rem;
        color: #64748b;
        font-weight: 400;
        letter-spacing: 0.05em;
        text-transform: uppercase;
    }
    .plasmart-divider {
        width: 1px; height: 48px;
        background: rgba(255,255,255,0.1);
        margin: 0 8px;
    }
    .stat-card {
        background: #1e2130;
        border: 1px solid rgba(255,255,255,0.07);
        border-radius: 14px;
        padding: 20px 24px;
        position: relative;
        overflow: hidden;
    }
    .stat-card::before {
        content: '';
        position: absolute;
        top: 0; left: 0;
        width: 3px; height: 100%;
        border-radius: 4px 0 0 4px;
    }
    .stat-card.orange::before { background: #f97316; }
    .stat-card.green::before  { background: #22c55e; }
    .stat-card.blue::before   { background: #60a5fa; }
    .stat-card.red::before    { background: #ef4444; }
    .stat-value {
        font-size: 1.8rem;
        font-weight: 700;
        font-family: 'JetBrains Mono', monospace;
        margin: 4px 0;
    }
    .stat-label {
        font-size: 0.75rem;
        color: #64748b;
        text-transform: uppercase;
        letter-spacing: 0.08em;
        font-weight: 500;
    }
    .stat-card.orange .stat-value { color: #f97316; }
    .stat-card.green .stat-value  { color: #22c55e; }
    .stat-card.blue .stat-value   { color: #60a5fa; }
    .stat-card.red .stat-value    { color: #ef4444; }
    .form-section {
        background: #1e2130;
        border: 1px solid rgba(255,255,255,0.07);
        border-radius: 14px;
        padding: 24px 28px;
        margin-bottom: 20px;
    }
    .form-section-title {
        font-size: 0.75rem;
        font-weight: 600;
        text-transform: uppercase;
        letter-spacing: 0.12em;
        color: #f97316;
        margin-bottom: 16px;
        display: flex;
        align-items: center;
        gap: 8px;
    }
    .calculo-box {
        background: linear-gradient(135deg, #1a1d27 0%, #1e2130 100%);
        border: 1px solid rgba(249,115,22,0.3);
        border-radius: 14px;
        padding: 24px;
        margin-top: 8px;
    }
    .calculo-row {
        display: flex;
        justify-content: space-between;
        align-items: center;
        padding: 10px 0;
        border-bottom: 1px solid rgba(255,255,255,0.07);
        font-size: 0.95rem;
    }
    .calculo-row:last-child { border-bottom: none; }
    .calculo-row.total {
        font-weight: 700;
        font-size: 1.1rem;
        color: #f97316;
        border-top: 2px solid rgba(249,115,22,0.3);
        margin-top: 4px;
        padding-top: 14px;
        border-bottom: none;
    }
    .calculo-label { color: #94a3b8; }
    .calculo-value { font-family: 'JetBrains Mono', monospace; font-weight: 600; }
    .ot-number {
        font-family: 'JetBrains Mono', monospace;
        font-weight: 700;
        color: #f97316;
        font-size: 1.2rem;
    }
    .badge {
        display: inline-block;
        padding: 3px 10px;
        border-radius: 20px;
        font-size: 0.72rem;
        font-weight: 600;
        letter-spacing: 0.05em;
        text-transform: uppercase;
    }
    .badge-abierta {
        background: rgba(34,197,94,0.15);
        color: #22c55e;
        border: 1px solid rgba(34,197,94,0.3);
    }
    .badge-cerrada {
        background: rgba(239,68,68,0.12);
        color: #f87171;
        border: 1px solid rgba(239,68,68,0.25);
    }
    .stTabs [data-baseweb="tab-list"] {
        background: #1e2130 !important;
        border-radius: 12px !important;
        padding: 6px !important;
        gap: 4px !important;
        border: 1px solid rgba(255,255,255,0.07) !important;
    }
    .stTabs [data-baseweb="tab"] {
        background: transparent !important;
        border-radius: 8px !important;
        color: #64748b !important;
        font-family: 'Sora', sans-serif !important;
        font-size: 0.9rem !important;
        font-weight: 500 !important;
        padding: 10px 22px !important;
    }
    .stTabs [aria-selected="true"] {
        background: #f97316 !important;
        color: white !important;
    }
    .stTabs [data-baseweb="tab-border"] { display: none !important; }
    .stButton > button {
        border-radius: 8px !important;
        font-family: 'Sora', sans-serif !important;
        font-weight: 600 !important;
        font-size: 0.9rem !important;
        transition: all 0.2s !important;
    }
    .stButton > button[kind="primary"] {
        background: #f97316 !important;
        border: none !important;
        color: white !important;
        box-shadow: 0 4px 16px rgba(249,115,22,0.35) !important;
    }
    .stButton > button[kind="primary"]:hover {
        background: #ea6c10 !important;
        transform: translateY(-1px) !important;
    }
    #MainMenu, footer, header { visibility: hidden; }
    .block-container { padding-top: 24px !important; padding-bottom: 40px !important; }
    hr { border-color: rgba(255,255,255,0.07) !important; margin: 24px 0 !important; }
</style>
""", unsafe_allow_html=True)

# ---------------------------------------------------------------------------
# CONSTANTES
# ---------------------------------------------------------------------------
VENDEDORES   = ["Marcelo", "Santi", "Agus"]
ORIGENES     = ["Redes", "Web", "Referencia"]
MEDIOS_PAGO  = ["Efectivo", "Transferencia"]
FACTURA_OPS  = ["Sí", "No"]
SCOPES       = ["https://spreadsheets.google.com/feeds",
                "https://www.googleapis.com/auth/drive"]

COLUMNAS = [
    "numero_ot", "fecha", "cliente", "origen_venta", "medio_pago",
    "vendedor", "kg_chapa", "con_factura", "valor_sin_iva",
    "iva", "total_venta", "anticipo_pct", "monto_anticipo",
    "saldo", "fecha_pago_anticipo", "fecha_entrega", "estado"
]

COLUMNAS_DISPLAY = {
    "numero_ot": "OT", "fecha": "Fecha", "cliente": "Cliente",
    "origen_venta": "Origen", "medio_pago": "Medio Pago", "vendedor": "Vendedor",
    "kg_chapa": "KG Chapa", "con_factura": "Factura", "valor_sin_iva": "Valor s/IVA",
    "iva": "IVA", "total_venta": "Total Venta", "anticipo_pct": "Anticipo %",
    "monto_anticipo": "Monto Anticipo", "saldo": "Saldo",
    "fecha_pago_anticipo": "F. Pago Anticipo", "fecha_entrega": "F. Entrega",
    "estado": "Estado",
}

# ---------------------------------------------------------------------------
# CONEXIÓN A GOOGLE SHEETS
# ---------------------------------------------------------------------------

@st.cache_resource(show_spinner=False)
def get_google_client():
    """Crea y cachea el cliente de Google Sheets autenticado."""
    creds_json = os.environ.get("GOOGLE_CREDS_JSON")
    if not creds_json:
        st.error("⛔ Variable de entorno GOOGLE_CREDS_JSON no encontrada.")
        st.stop()
    try:
        creds_dict = json.loads(creds_json)
    except json.JSONDecodeError:
        st.error("⛔ GOOGLE_CREDS_JSON no es un JSON válido. Verificá que esté bien pegado en Railway.")
        st.stop()
    # Railway a veces escapa los \n del private_key — los restauramos
    if "private_key" in creds_dict:
        creds_dict["private_key"] = creds_dict["private_key"].replace("\\n", "\n")
    creds = Credentials.from_service_account_info(creds_dict, scopes=SCOPES)
    return gspread.authorize(creds)


def get_sheet():
    """Retorna el worksheet principal, creándolo si no existe."""
    client = get_google_client()
    sheet_id = os.environ.get("GOOGLE_SHEET_ID")
    if not sheet_id:
        st.error("⛔ Variable de entorno GOOGLE_SHEET_ID no encontrada.")
        st.stop()
    spreadsheet = client.open_by_key(sheet_id)
    try:
        ws = spreadsheet.worksheet("Órdenes")
    except gspread.WorksheetNotFound:
        ws = spreadsheet.add_worksheet(title="Órdenes", rows=1000, cols=len(COLUMNAS))
        ws.append_row(COLUMNAS)
    return ws


# ---------------------------------------------------------------------------
# OPERACIONES DE DATOS
# ---------------------------------------------------------------------------

def leer_ordenes() -> pd.DataFrame:
    """Lee todas las OTs del Google Sheet y retorna un DataFrame."""
    try:
        ws = get_sheet()
        data = ws.get_all_records(expected_headers=COLUMNAS)
        if not data:
            return pd.DataFrame(columns=COLUMNAS)
        df = pd.DataFrame(data)
        # Asegurar columnas
        for col in COLUMNAS:
            if col not in df.columns:
                df[col] = None
        # Convertir tipos
        for col_fecha in ["fecha", "fecha_pago_anticipo", "fecha_entrega"]:
            df[col_fecha] = pd.to_datetime(df[col_fecha], errors="coerce")
        for col_num in ["kg_chapa", "valor_sin_iva", "iva", "total_venta",
                        "monto_anticipo", "saldo", "anticipo_pct"]:
            df[col_num] = pd.to_numeric(df[col_num], errors="coerce")
        df["numero_ot"] = df["numero_ot"].astype(str)
        return df[COLUMNAS]
    except Exception as e:
        st.error(f"Error leyendo Google Sheets: {e}")
        return pd.DataFrame(columns=COLUMNAS)


def agregar_ot(nueva_ot: dict):
    """Agrega una fila nueva al final del Sheet."""
    ws = get_sheet()
    fila = [str(nueva_ot.get(col, "")) for col in COLUMNAS]
    ws.append_row(fila, value_input_option="USER_ENTERED")


def actualizar_fila(numero_ot: str, datos: dict):
    """Actualiza la fila completa de una OT existente buscando por numero_ot."""
    ws = get_sheet()
    # Buscar la celda con el número de OT en la columna A
    cell = ws.find(numero_ot, in_column=1)
    if not cell:
        st.error(f"No se encontró la OT {numero_ot} en el Sheet.")
        return
    fila_valores = [str(datos.get(col, "")) for col in COLUMNAS]
    rango = f"A{cell.row}:{chr(64 + len(COLUMNAS))}{cell.row}"
    ws.update(rango, [fila_valores], value_input_option="USER_ENTERED")


def cerrar_ot(numero_ot: str):
    """Cambia el estado de una OT a Cerrada."""
    ws = get_sheet()
    cell = ws.find(numero_ot, in_column=1)
    if not cell:
        st.error(f"No se encontró la OT {numero_ot}.")
        return
    col_estado = COLUMNAS.index("estado") + 1
    ws.update_cell(cell.row, col_estado, "Cerrada")


def generar_numero_ot(df: pd.DataFrame) -> str:
    """Genera el siguiente número de OT correlativo del año actual."""
    anio = datetime.now().year
    prefijo = f"OT-{anio}-"
    if df.empty:
        return f"{prefijo}0001"
    ots_anio = df[df["numero_ot"].str.startswith(prefijo, na=False)]["numero_ot"]
    if ots_anio.empty:
        return f"{prefijo}0001"
    numeros = ots_anio.str.replace(prefijo, "", regex=False).str.extract(r"(\d+)")[0].dropna().astype(int)
    siguiente = numeros.max() + 1 if not numeros.empty else 1
    return f"{prefijo}{siguiente:04d}"


def calcular_valores(valor_sin_iva: float, con_factura: bool, anticipo_pct: int):
    iva = round(valor_sin_iva * 0.21, 2) if con_factura else 0.0
    total = round(valor_sin_iva + iva, 2)
    monto_anticipo = round(total * anticipo_pct / 100, 2)
    saldo = round(total - monto_anticipo, 2)
    return iva, total, monto_anticipo, saldo


def exportar_excel_bytes(df: pd.DataFrame) -> bytes:
    """Genera un archivo Excel en memoria para descargar."""
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    output = io.BytesIO()
    df_exp = df.copy()
    for col_fecha in ["fecha", "fecha_pago_anticipo", "fecha_entrega"]:
        df_exp[col_fecha] = pd.to_datetime(df_exp[col_fecha], errors="coerce").dt.strftime("%d/%m/%Y")
    df_exp.rename(columns=COLUMNAS_DISPLAY, inplace=True)
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_exp.to_excel(writer, sheet_name="Órdenes", index=False)
        ws = writer.sheets["Órdenes"]
        fill = PatternFill(start_color="1E2130", end_color="1E2130", fill_type="solid")
        font = Font(color="F97316", bold=True, size=11)
        for col_num in range(1, len(COLUMNAS) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(col_num)].width = 18
    return output.getvalue()


# ---------------------------------------------------------------------------
# HEADER
# ---------------------------------------------------------------------------
st.markdown(f"""
<div class="plasmart-header">
    <div>
        <div class="plasmart-logo">Plas<span>mart</span></div>
        <div class="plasmart-subtitle">plasmartcba.com</div>
    </div>
    <div class="plasmart-divider"></div>
    <div style="flex:1">
        <div style="font-size:1.1rem;font-weight:600;color:#f1f5f9">Gestor de Órdenes de Trabajo</div>
        <div style="font-size:0.82rem;color:#64748b;margin-top:2px">Sistema interno · Chapa y metalúrgica</div>
    </div>
    <div style="text-align:right;font-family:'JetBrains Mono',monospace;font-size:0.8rem;color:#64748b">
        {datetime.now().strftime("%d/%m/%Y · %H:%M")}
    </div>
</div>
""", unsafe_allow_html=True)

# ---------------------------------------------------------------------------
# CARGA INICIAL DE DATOS
# ---------------------------------------------------------------------------
# Usamos session_state para cachear los datos y refrescarlos con un botón
if "df" not in st.session_state or st.session_state.get("forzar_refresh"):
    with st.spinner("Conectando con Google Sheets…"):
        st.session_state.df = leer_ordenes()
    st.session_state.forzar_refresh = False

df_global = st.session_state.df

# ---------------------------------------------------------------------------
# MÉTRICAS
# ---------------------------------------------------------------------------
total_ots       = len(df_global)
abiertas        = len(df_global[df_global["estado"] == "Abierta"]) if not df_global.empty else 0
cerradas        = len(df_global[df_global["estado"] == "Cerrada"]) if not df_global.empty else 0
total_facturado = df_global["total_venta"].sum() if not df_global.empty else 0

col_m1, col_m2, col_m3, col_m4 = st.columns(4)
with col_m1:
    st.markdown(f'<div class="stat-card orange"><div class="stat-label">Total OTs</div><div class="stat-value">{total_ots}</div></div>', unsafe_allow_html=True)
with col_m2:
    st.markdown(f'<div class="stat-card green"><div class="stat-label">Abiertas</div><div class="stat-value">{abiertas}</div></div>', unsafe_allow_html=True)
with col_m3:
    st.markdown(f'<div class="stat-card red"><div class="stat-label">Cerradas</div><div class="stat-value">{cerradas}</div></div>', unsafe_allow_html=True)
with col_m4:
    st.markdown(f'<div class="stat-card blue"><div class="stat-label">Total Facturado</div><div class="stat-value">${total_facturado:,.0f}</div></div>', unsafe_allow_html=True)

st.markdown("<br>", unsafe_allow_html=True)

# Botón de actualizar datos en tiempo real
col_ref, _ = st.columns([1, 5])
with col_ref:
    if st.button("🔄 Actualizar datos", use_container_width=True):
        st.session_state.forzar_refresh = True
        st.rerun()

st.markdown("<br>", unsafe_allow_html=True)

# ---------------------------------------------------------------------------
# TABS
# ---------------------------------------------------------------------------
tab1, tab2 = st.tabs(["➕  Cargar Nueva OT", "📋  Ver y Gestionar Órdenes"])


# ===========================================================================
# TAB 1 — NUEVA OT
# ===========================================================================
with tab1:
    proximo_ot = generar_numero_ot(df_global)

    st.markdown(f"""
    <div style="display:flex;align-items:center;gap:12px;margin-bottom:24px">
        <div style="color:#64748b;font-size:0.85rem;font-weight:500;text-transform:uppercase;letter-spacing:0.08em">
            Próxima OT asignada:
        </div>
        <div class="ot-number">{proximo_ot}</div>
    </div>
    """, unsafe_allow_html=True)

    col_form, col_calc = st.columns([3, 2], gap="large")

    with col_form:
        st.markdown('<div class="form-section">', unsafe_allow_html=True)
        st.markdown('<div class="form-section-title">📁 Información General</div>', unsafe_allow_html=True)
        col_f1, col_f2 = st.columns(2)
        with col_f1:
            fecha = st.date_input("Fecha", value=date.today(), key="nf_fecha")
        with col_f2:
            vendedor = st.selectbox("Vendedor", VENDEDORES, key="nf_vendedor")
        cliente = st.text_input("Cliente", placeholder="Nombre del cliente…", key="nf_cliente")
        col_f3, col_f4 = st.columns(2)
        with col_f3:
            origen = st.selectbox("Origen de la Venta", ORIGENES, key="nf_origen")
        with col_f4:
            medio_pago = st.selectbox("Medio de Pago", MEDIOS_PAGO, key="nf_medio")
        st.markdown('</div>', unsafe_allow_html=True)

        st.markdown('<div class="form-section">', unsafe_allow_html=True)
        st.markdown('<div class="form-section-title">💰 Datos Económicos</div>', unsafe_allow_html=True)
        col_e1, col_e2 = st.columns(2)
        with col_e1:
            kg_chapa = st.number_input("KG de Chapa", min_value=0.0, step=1.0, format="%.1f", key="nf_kg")
        with col_e2:
            con_factura_str = st.selectbox("Con Factura", FACTURA_OPS, key="nf_factura")
            con_factura = con_factura_str == "Sí"
        valor_sin_iva = st.number_input(
            "Valor de la Venta (SIN IVA) $", min_value=0.0, step=1000.0, format="%.2f", key="nf_valor"
        )
        anticipo_pct = st.slider(
            "Anticipo (%)", min_value=0, max_value=100, step=5, value=50, key="nf_anticipo"
        )
        col_e3, col_e4 = st.columns(2)
        with col_e3:
            fecha_pago_ant = st.date_input("Fecha de pago del anticipo", value=date.today(), key="nf_fecha_ant")
        with col_e4:
            fecha_entrega = st.date_input("Fecha teórica de entrega", value=date.today(), key="nf_fecha_ent")
        st.markdown('</div>', unsafe_allow_html=True)

    with col_calc:
        iva, total_venta, monto_anticipo, saldo = calcular_valores(valor_sin_iva, con_factura, anticipo_pct)

        st.markdown("##### 🧮 Resumen en tiempo real")
        st.markdown(f"""
        <div class="calculo-box">
            <div class="calculo-row">
                <span class="calculo-label">Valor SIN IVA</span>
                <span class="calculo-value">${valor_sin_iva:,.2f}</span>
            </div>
            <div class="calculo-row">
                <span class="calculo-label">IVA ({"21%" if con_factura else "0% · sin factura"})</span>
                <span class="calculo-value">${iva:,.2f}</span>
            </div>
            <div class="calculo-row total">
                <span class="calculo-label">TOTAL VENTA</span>
                <span class="calculo-value">${total_venta:,.2f}</span>
            </div>
            <div class="calculo-row">
                <span class="calculo-label">Anticipo ({anticipo_pct}%)</span>
                <span class="calculo-value">${monto_anticipo:,.2f}</span>
            </div>
            <div class="calculo-row">
                <span class="calculo-label" style="color:#f87171">Saldo a cobrar</span>
                <span class="calculo-value" style="color:#f87171">${saldo:,.2f}</span>
            </div>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown("""
        <div style="background:rgba(249,115,22,0.08);border:1px solid rgba(249,115,22,0.2);
                    border-radius:10px;padding:14px 16px;font-size:0.82rem;color:#94a3b8">
            <strong style="color:#f97316">ℹ️ Guardado en la nube</strong><br>
            La OT se guarda directamente en Google Sheets y queda disponible
            para todos los vendedores en tiempo real.
        </div>
        """, unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)

        if st.button("💾  Guardar Nueva OT", type="primary", use_container_width=True, key="btn_guardar"):
            errores = []
            if not cliente.strip():
                errores.append("El campo **Cliente** es obligatorio.")
            if valor_sin_iva <= 0:
                errores.append("El **Valor de la Venta** debe ser mayor a 0.")
            if kg_chapa <= 0:
                errores.append("Los **KG de Chapa** deben ser mayores a 0.")

            if errores:
                for e in errores:
                    st.error(e)
            else:
                # Refrescar para número correlativo actualizado
                df_fresco = leer_ordenes()
                num_ot = generar_numero_ot(df_fresco)

                nueva_ot = {
                    "numero_ot":          num_ot,
                    "fecha":              fecha.strftime("%d/%m/%Y"),
                    "cliente":            cliente.strip(),
                    "origen_venta":       origen,
                    "medio_pago":         medio_pago,
                    "vendedor":           vendedor,
                    "kg_chapa":           kg_chapa,
                    "con_factura":        con_factura_str,
                    "valor_sin_iva":      valor_sin_iva,
                    "iva":                iva,
                    "total_venta":        total_venta,
                    "anticipo_pct":       anticipo_pct,
                    "monto_anticipo":     monto_anticipo,
                    "saldo":              saldo,
                    "fecha_pago_anticipo": fecha_pago_ant.strftime("%d/%m/%Y"),
                    "fecha_entrega":      fecha_entrega.strftime("%d/%m/%Y"),
                    "estado":             "Abierta",
                }

                with st.spinner("Guardando en Google Sheets…"):
                    agregar_ot(nueva_ot)

                st.success(f"✅  OT **{num_ot}** cargada correctamente para **{cliente.strip()}**.")
                st.balloons()
                st.session_state.forzar_refresh = True
                time.sleep(1)
                st.rerun()


# ===========================================================================
# TAB 2 — VER Y GESTIONAR
# ===========================================================================
with tab2:
    df = st.session_state.df.copy()

    # ── Filtros ─────────────────────────────────────────────────────────
    st.markdown('<div class="form-section">', unsafe_allow_html=True)
    st.markdown('<div class="form-section-title">🔍 Filtros</div>', unsafe_allow_html=True)
    col_fil1, col_fil2, col_fil3, col_fil4, col_fil5 = st.columns(5)
    with col_fil1:
        f_estado = st.selectbox("Estado", ["Todos", "Abierta", "Cerrada"], key="fil_estado")
    with col_fil2:
        f_vendedor = st.selectbox("Vendedor", ["Todos"] + VENDEDORES, key="fil_vendedor")
    with col_fil3:
        f_origen = st.selectbox("Origen", ["Todos"] + ORIGENES, key="fil_origen")
    with col_fil4:
        f_cliente = st.text_input("Cliente", placeholder="Buscar…", key="fil_cliente")
    with col_fil5:
        f_fecha_desde = st.date_input("Desde", value=None, key="fil_desde")
    st.markdown('</div>', unsafe_allow_html=True)

    df_filtrado = df.copy()
    if f_estado != "Todos":
        df_filtrado = df_filtrado[df_filtrado["estado"] == f_estado]
    if f_vendedor != "Todos":
        df_filtrado = df_filtrado[df_filtrado["vendedor"] == f_vendedor]
    if f_origen != "Todos":
        df_filtrado = df_filtrado[df_filtrado["origen_venta"] == f_origen]
    if f_cliente.strip():
        df_filtrado = df_filtrado[df_filtrado["cliente"].str.contains(f_cliente.strip(), case=False, na=False)]
    if f_fecha_desde:
        df_filtrado = df_filtrado[df_filtrado["fecha"] >= pd.Timestamp(f_fecha_desde)]

    # ── Barra de acciones ────────────────────────────────────────────────
    col_info, col_export = st.columns([3, 1])
    with col_info:
        st.markdown(f"""
        <div style="color:#94a3b8;font-size:0.9rem;padding:8px 0">
            Mostrando <strong style="color:#f1f5f9">{len(df_filtrado)}</strong> órdenes
            {"de " + str(len(df)) + " totales" if len(df_filtrado) != len(df) else "en total"}
        </div>
        """, unsafe_allow_html=True)
    with col_export:
        if not df_filtrado.empty:
            excel_bytes = exportar_excel_bytes(df_filtrado)
            st.download_button(
                label="⬇  Descargar Excel",
                data=excel_bytes,
                file_name=f"plasmart_OTs_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="btn_export"
            )

    # ── Tabla ────────────────────────────────────────────────────────────
    if df_filtrado.empty:
        st.info("No se encontraron órdenes con los filtros seleccionados.")
    else:
        df_display = df_filtrado.copy()
        for col_fecha in ["fecha", "fecha_pago_anticipo", "fecha_entrega"]:
            df_display[col_fecha] = pd.to_datetime(df_display[col_fecha], errors="coerce").dt.strftime("%d/%m/%Y")
        for col_num in ["valor_sin_iva", "iva", "total_venta", "monto_anticipo", "saldo"]:
            df_display[col_num] = df_display[col_num].apply(lambda x: f"${x:,.2f}" if pd.notna(x) else "-")
        df_display["anticipo_pct"] = df_display["anticipo_pct"].apply(lambda x: f"{int(x)}%" if pd.notna(x) else "-")
        df_display.rename(columns=COLUMNAS_DISPLAY, inplace=True)

        st.dataframe(df_display, use_container_width=True, hide_index=True, height=360)

        st.markdown("---")

        # ── Editar / Cerrar ──────────────────────────────────────────────
        st.markdown("#### ✏️ Editar / Cerrar Orden")
        ots_disponibles = df_filtrado["numero_ot"].tolist()
        ot_sel = st.selectbox("Seleccionar OT", ots_disponibles, key="sel_ot")

        if ot_sel:
            ot_data = df[df["numero_ot"] == ot_sel].iloc[0]
            estado_actual = ot_data["estado"]

            st.markdown(f"""
            <div style="display:flex;align-items:center;gap:12px;margin:12px 0">
                <span class="ot-number">{ot_sel}</span>
                <span class="badge badge-{'abierta' if estado_actual == 'Abierta' else 'cerrada'}">{estado_actual}</span>
                <span style="color:#64748b;font-size:0.85rem">· {ot_data.get('cliente','')}</span>
            </div>
            """, unsafe_allow_html=True)

            with st.expander("✏️ Editar esta orden", expanded=False):
                col_ed1, col_ed2 = st.columns(2)
                with col_ed1:
                    try:
                        val_fecha = pd.to_datetime(ot_data["fecha"], dayfirst=True).date()
                    except Exception:
                        val_fecha = date.today()
                    ed_fecha    = st.date_input("Fecha", value=val_fecha, key="ed_fecha")
                    ed_cliente  = st.text_input("Cliente", value=str(ot_data["cliente"]), key="ed_cliente")
                    ed_origen   = st.selectbox("Origen", ORIGENES,
                                               index=ORIGENES.index(ot_data["origen_venta"]) if ot_data["origen_venta"] in ORIGENES else 0,
                                               key="ed_origen")
                    ed_medio    = st.selectbox("Medio de Pago", MEDIOS_PAGO,
                                               index=MEDIOS_PAGO.index(ot_data["medio_pago"]) if ot_data["medio_pago"] in MEDIOS_PAGO else 0,
                                               key="ed_medio")
                    ed_vendedor = st.selectbox("Vendedor", VENDEDORES,
                                               index=VENDEDORES.index(ot_data["vendedor"]) if ot_data["vendedor"] in VENDEDORES else 0,
                                               key="ed_vendedor")
                with col_ed2:
                    ed_kg      = st.number_input("KG Chapa", value=float(ot_data["kg_chapa"] or 0), step=1.0, key="ed_kg")
                    ed_factura = st.selectbox("Con Factura", FACTURA_OPS,
                                              index=FACTURA_OPS.index(ot_data["con_factura"]) if ot_data["con_factura"] in FACTURA_OPS else 0,
                                              key="ed_factura")
                    ed_valor   = st.number_input("Valor SIN IVA $", value=float(ot_data["valor_sin_iva"] or 0), step=1000.0, key="ed_valor")
                    ed_ant_pct = st.slider("Anticipo (%)", 0, 100, int(ot_data["anticipo_pct"] or 0), step=5, key="ed_ant_pct")
                    try:
                        val_fecha_ant = pd.to_datetime(ot_data["fecha_pago_anticipo"], dayfirst=True).date()
                    except Exception:
                        val_fecha_ant = date.today()
                    try:
                        val_fecha_ent = pd.to_datetime(ot_data["fecha_entrega"], dayfirst=True).date()
                    except Exception:
                        val_fecha_ent = date.today()
                    ed_fecha_ant = st.date_input("F. Pago Anticipo", value=val_fecha_ant, key="ed_fecha_ant")
                    ed_fecha_ent = st.date_input("F. Entrega", value=val_fecha_ent, key="ed_fecha_ent")

                ed_iva, ed_total, ed_monto_ant, ed_saldo = calcular_valores(ed_valor, ed_factura == "Sí", ed_ant_pct)

                st.markdown(f"""
                <div class="calculo-box" style="margin-top:12px">
                    <div class="calculo-row">
                        <span class="calculo-label">IVA</span>
                        <span class="calculo-value">${ed_iva:,.2f}</span>
                    </div>
                    <div class="calculo-row total">
                        <span class="calculo-label">Total Venta</span>
                        <span class="calculo-value">${ed_total:,.2f}</span>
                    </div>
                    <div class="calculo-row">
                        <span class="calculo-label">Anticipo ({ed_ant_pct}%)</span>
                        <span class="calculo-value">${ed_monto_ant:,.2f}</span>
                    </div>
                    <div class="calculo-row">
                        <span class="calculo-label" style="color:#f87171">Saldo</span>
                        <span class="calculo-value" style="color:#f87171">${ed_saldo:,.2f}</span>
                    </div>
                </div>
                """, unsafe_allow_html=True)

                if st.button("💾 Guardar Cambios", type="primary", key="btn_edit_save"):
                    if not ed_cliente.strip():
                        st.error("El campo Cliente es obligatorio.")
                    else:
                        datos_actualizados = {
                            "numero_ot":           ot_sel,
                            "fecha":               ed_fecha.strftime("%d/%m/%Y"),
                            "cliente":             ed_cliente.strip(),
                            "origen_venta":        ed_origen,
                            "medio_pago":          ed_medio,
                            "vendedor":            ed_vendedor,
                            "kg_chapa":            ed_kg,
                            "con_factura":         ed_factura,
                            "valor_sin_iva":       ed_valor,
                            "iva":                 ed_iva,
                            "total_venta":         ed_total,
                            "anticipo_pct":        ed_ant_pct,
                            "monto_anticipo":      ed_monto_ant,
                            "saldo":               ed_saldo,
                            "fecha_pago_anticipo": ed_fecha_ant.strftime("%d/%m/%Y"),
                            "fecha_entrega":       ed_fecha_ent.strftime("%d/%m/%Y"),
                            "estado":              estado_actual,
                        }
                        with st.spinner("Guardando en Google Sheets…"):
                            actualizar_fila(ot_sel, datos_actualizados)
                        st.success(f"✅ OT {ot_sel} actualizada correctamente.")
                        st.session_state.forzar_refresh = True
                        time.sleep(1)
                        st.rerun()

            # ── Cerrar orden ─────────────────────────────────────────────
            if estado_actual == "Abierta":
                st.markdown("<br>", unsafe_allow_html=True)
                st.markdown("""
                <div style="background:rgba(239,68,68,0.08);border:1px solid rgba(239,68,68,0.2);
                            border-radius:10px;padding:14px 18px;margin-bottom:12px;font-size:0.85rem;color:#94a3b8">
                    ⚠️ <strong style="color:#f87171">Atención</strong> — Esta acción cambia el estado
                    a <strong>Cerrada</strong> de forma permanente.
                </div>
                """, unsafe_allow_html=True)
                confirmar = st.checkbox(f"Confirmo que deseo cerrar la orden {ot_sel}", key="check_cerrar")
                if st.button(f"🔒  Cerrar Orden {ot_sel}", disabled=not confirmar,
                             key="btn_cerrar", use_container_width=True):
                    with st.spinner("Cerrando orden en Google Sheets…"):
                        cerrar_ot(ot_sel)
                    st.success(f"✅ Orden **{ot_sel}** cerrada correctamente.")
                    st.session_state.forzar_refresh = True
                    time.sleep(1)
                    st.rerun()
            else:
                st.markdown("""
                <div style="background:rgba(239,68,68,0.06);border:1px solid rgba(239,68,68,0.15);
                            border-radius:10px;padding:12px 16px;font-size:0.85rem;color:#64748b">
                    🔒 Esta orden ya se encuentra <strong style="color:#f87171">Cerrada</strong>.
                </div>
                """, unsafe_allow_html=True)
